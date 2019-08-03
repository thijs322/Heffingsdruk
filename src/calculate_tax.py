#!python3
import urllib.request
import re
import time
import string
import datetime
import openpyxl
from pathlib import Path
import os
import matplotlib.pyplot as plt
import ssl

'''
TODO arbeidsinkomen berekening niet hardcoden
TODO roken/alcohol accijns
TODO Energietaks = stroomverbruik * >>STROOMTAKS<< + gasverbruik * >>GASTAKS<<

VOORBEELD
boris = Persoon('5531vg',25,2500,0,300,600,2000,35000,1500,3000,93)
auto = Voertuig(boris,'85tdpv',2500,20000)
auto.get_data()
tax = Belasting(boris,auto,2,0)
resultaat = Calculation(boris,auto,tax)
'''


class Persoon:

    def __init__(
        self, postcode: str, leeftijd:int, bruto_loon_mnd:int, bonus:int, uitgaven_laag:int, uitgaven_hoog:int,
        spaargeld:int, schulden:int, verbruik_gas:int, verbruik_stroom:int, verbruik_water:int):

        self.postcode = postcode
        self.leeftijd = leeftijd
        self.bruto_loon_mnd = bruto_loon_mnd
        self.bonus = bonus
        self.uitgaven_laag = uitgaven_laag * 12
        self.uitgaven_hoog = uitgaven_hoog * 12
        self.spaargeld = spaargeld
        self.schulden = schulden
        self.verbruik_gas = verbruik_gas
        self.verbruik_stroom = verbruik_stroom
        self.verbruik_water = verbruik_water
        self.bruto_loon_jr = self.bruto_loon_mnd * 12
        self.vakantiegeld = self.bruto_loon_jr * 0.08
        self.roken = 0
        self.alcohol = 0
        self.loon_speciaal = self.vakantiegeld + self.bonus
        self.loon_totaal = self.bruto_loon_jr + self.loon_speciaal

        '''
        Informatie over de woonplaats van de persoon (stad, gemeente, provincie) 
        wordt automatisch opgehaald aan de hand van de postcode
        '''
        print('Getting location data ...')
        start = time.perf_counter()
        url = 'https://www.postcodezoekmachine.nl/' + self.postcode.upper()
        response = urllib.request.urlopen(url)
        html = response.read().decode(response.headers.get_content_charset())

        self.stad = regex_lookup("<td><a href=\"\/(\w*)\">", html)
        self.stad = string.capwords(self.stad, '-')
        self.provincie = regex_lookup("<td><a href=\"\/provincie\/(.*?)\">", html)
        self.provincie = string.capwords(self.provincie, '-')
        self.gemeente = regex_lookup("<td><a href=\"\/gemeente\/(.*?)\">", html)
        self.gemeente = string.capwords(self.gemeente, '-')
        print('Success! Data found in {} seconds'.format(round(time.perf_counter() - start, 2)))
        print('{}, {}, {}'.format(self.stad, self.provincie, self.gemeente))

    def roken(self, roken_per_week):
        self.roken = roken_per_week

    def drinken(self, alcohol_per_week):
        self.alcohol = alcohol_per_week


class Voertuig:

    def __init__(self, persoon:str, kenteken:int, prijs:int, km_jaar:int):
        self.persoon = persoon
        self.kenteken = kenteken
        self.prijs = prijs
        self.km_jaar = km_jaar

        # Zorg ervoor dat het kenteken 6 of 8 (met streepjes ertussen) tekens lang is
        if not (len(self.kenteken) != 8 or len(self.kenteken) != 6):
            self.kenteken = input('Kenteken moet 6 (12abcd) of 8 (12-ab-cd) tekens zijn >>> ')

        """
        Zet kenteken in de goede vorm (12-AB-CD of 12-ABC-3) ongeacht van input
        Altijd een streepje tussen letter en cijfer

        probeer str[0] en int[1] of andersom
        zo ja>streepje ertussen
        zo nee > i+1
        """
        if len(self.kenteken) == 6:
            dash_idx = []
            kenteken_temp = []

            # Kijk op welke plek een str of een int zit
            for i in range(len(self.kenteken)):
                try:
                    int(self.kenteken[i])
                    dash_idx.append(1)
                except ValueError:
                    dash_idx.append(0)

            # Zet streepjes tussen elke twee opeenvolgende verschillende tekens
            for i in range(1, len(dash_idx)):
                kenteken_temp.append(self.kenteken[i - 1])
                if dash_idx[i] != dash_idx[i - 1]:
                    kenteken_temp.append('-')
            kenteken_temp.append(self.kenteken[-1])  # Voeg laatste teken toe
            self.kenteken = ''.join(kenteken_temp)

            # Als er maar 1 streepje in zit, zet een extra streepje tussen 4 opeenvolgende tekens
            if self.kenteken.count('-') == 1:
                if sum(dash_idx[:4]) == 0 or sum(dash_idx[:4]) == 4:  # first four are the same
                    self.kenteken = self.kenteken[:2] + '-' + self.kenteken[2:]
                else:  # last four are the same
                    self.kenteken = self.kenteken[:5] + '-' + self.kenteken[5:]

        self.kenteken = self.kenteken.upper()

        '''
        Haal automatisch informatie over de auto op
        '''
        print('Getting vehicle data ...')
        start = time.perf_counter()

        # Get html data
        ini_url = 'https://www.kentekencheck.nl/kenteken?i=' + self.kenteken
        ini_html = read_html(ini_url)
        time.sleep(1)
        link_to_rapport = regex_lookup("<iframe src=\"(.*?)\"", ini_html)
        html = read_html('https://www.kentekencheck.nl' + link_to_rapport)

        # Read stuff
        self.merk_nummer = regex_lookup("Merk<\/td>\s*<td style=\"width:60%;\">(.*?)<\/td>", html) + ' ' + regex_lookup(
            "<td>Type<\/td>\s*<td>(.*?)<\/td>", html)
        self.bouwjaar = regex_lookup("<td>Bouwjaar<\/td>\s*<td>.*?(\d{4})<\/td>", html)
        self.brandstof = regex_lookup("<td>Brandstof<\/td>\s*<td>(.*?)<\/td>", html)
        self.nieuwprijs = regex_lookup("<td>Nieuwprijs<\/td>\s*<td>&euro; (.*?)<\/td>", html).replace('.', '')
        self.gewicht = regex_lookup("<td>Massa ledig voertuig<\/td>\s*<td>(\d*) KG<\/td>", html)
        self.wegenbelasting = int(
            regex_lookup("<td>" + self.persoon.provincie + "<\/td>\s*.*?<\/td>\s*<td>&euro;(\d*)<\/td>", html))
        try:
            self.verbruik = round(
                1 / float(regex_lookup("<td>Verbruik gecombineerd<\/td>\s*<td>.*\(1:(.*?)km\)<\/td>", html)), 4)
        except AttributeError:
            print('Verbruik niet gevonden')
        try:
            self.CO2_uitstoot = int(regex_lookup("<td>CO2 uitstoot<\/td>\s*<td>(\d*?) g\/km<\/td>", html))
        except AttributeError:
            print('CO2 uitstoot niet gevonden')

        print('Success! Data found in {} seconds'.format(round(time.perf_counter() - start, 2)))
        print('{}, {}, {}, {}, {}, {}, {}, {}'.format(self.merk_nummer, self.bouwjaar, self.brandstof, self.nieuwprijs,
                                                      self.gewicht, self.wegenbelasting, str(self.verbruik),
                                                      str(self.CO2_uitstoot)))


class Belasting:

    def __init__(self, persoon, voertuig, huishouden_personen, fiscale_partner):
        """
        Verschillende belastingen worden hier berekend op basis van persoons/voertuigs gegevens

        :param persoon:             naam van de persoon gedefiniëerd in class Persoon
        :param auto:                naam van het voertuig gedefiniëerd in class Voertuig
        :param huishouden_personen: aantal personen waarmee je samenwoont
        :param fiscale_partner:     afwezigheid (0) of aanwezigheid (1) van een fiscale partner
        """
        self.persoon = persoon
        self.auto = voertuig
        self.fiscale_partner = fiscale_partner
        try:
            self.huishouden_personen = int(huishouden_personen)
        except ValueError:
            self.huishouden_personen = input('Vul een geldig getal in voor het aantal personen in huis: >>> ')

        current_year = str(datetime.datetime.now().year)

        # Auto
        print('Getting car...')
        url = "https://www.unitedconsumers.com/tanken/informatie/opbouw-brandstofprijzen.asp"
        html = read_html(url)
        self.benzine_prijs = float(regex_lookup("Opbouw Benzine .*?<strong>(.*?)<\/strong>", html).replace(',', '.'))
        self.benzine_btw = int(regex_lookup("Opbouw Benzine .*?BTW.*?(\d*)%", html)) / 100
        self.benzine_accijns = int(regex_lookup("Opbouw Benzine .*?Accijns.*?(\d*)%", html)) / 100
        self.diesel_prijs = float(regex_lookup("Opbouw Diesel .*?<strong>(.*?)<\/strong>", html).replace(',', '.'))
        self.diesel_btw = int(regex_lookup("Opbouw Diesel .*?BTW.*?(\d*)%", html)) / 100
        self.diesel_accijns = int(regex_lookup("Opbouw Diesel .*?Accijns.*?(\d*)%", html)) / 100
        self.MRB = self.auto.wegenbelasting

        # Gemeente
        print('Getting gemeente...')
        url = "https://www.coelo.nl/images/Gemeentelijke_belastingen_{}.xlsx".format(current_year)
        url_name = url.split('/')[-1]
        url_opcenten = "https://www.coelo.nl/images/Provinciale_opcenten_{}.xlsx".format(current_year)
        url_opcenten_name = url_opcenten.split('/')[-1]

        # Download excel tabellen als ze nog niet bestaan
        if not os.path.abspath(url_name):
            urllib.request.urlretrieve(url, url_name)
        if not os.path.abspath(url_opcenten_name):
            urllib.request.urlretrieve(url_opcenten, url_opcenten_name)

        wb = openpyxl.load_workbook(Path(__file__).parent / url_name)['Gegevens per gemeente']
        if self.huishouden_personen > 1:
            col = ['I', 'N', 'S']
        else:
            col = ['I', 'L', 'Q']
        for row in range(5, wb.max_row + 1):
            if wb['B' + str(row)].value == self.persoon.gemeente:
                self.OZB = round(wb[col[0] + str(row)].value, 2)
                self.afvalheffing = round(wb[col[1] + str(row)].value, 2)
                self.rioolheffing = wb[col[2] + str(row)].value
                break

        # Opcenten
        wb = openpyxl.load_workbook(Path(__file__).parent / url_opcenten_name)['Gegevens per provincie']
        for row in range(6, 18):
            if wb['B' + str(row)].value == self.persoon.provincie:
                self.opcenten = round(wb['C' + str(row)].value / 100, 4)
                break

        # Loonschaal
        url = "https://www.belastingdienst.nl/bibliotheek/handboeken/html/boeken/HL/handboek_loonheffingen_2019-tarieven_bedragen_en_percentages.html"
        html = read_html(url)

        self.loonbelasting_schaal = [
            0,
            int(regex_lookup("rmkrnpakgd.*?t\/m € (.*?)<\/p>", html).replace('.', '')),
            int(regex_lookup("bdoeboonge.*?t\/m € (.*?)<\/p>", html).replace('.', '')),
            int(regex_lookup("eablhjemgh.*?t\/m € (.*?)<\/p>", html).replace('.', ''))
        ]
        self.loonbelasting = [
            float(regex_lookup("obcfqdbaga\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("ehdmneflgk\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("eqqaokdfgo\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("eoadoaopgf\">(.*?)%", html).replace(',', '.')) / 100,
        ]
        self.loonbelasting_aow = [
            float(regex_lookup("pdncrhorgl\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("bkmcoadagj\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("ldfkdkfqge\">(.*?)%", html).replace(',', '.')) / 100,
            float(regex_lookup("khpffjqjgj\">(.*?)%", html).replace(',', '.')) / 100
        ]

        # Vermogensbelasting
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/vermogen_en_aanmerkelijk_belang/vermogen/belasting_betalen_over_uw_vermogen/grondslag_sparen_en_beleggen/berekening-2019/berekening-belasting-over-inkomen-uit-vermogen-over-2019"
        html = read_html(url)

        self.vermogensbelasting_schaal = [
            0,
            int(regex_lookup("Tot en met €&nbsp;(.*?)<\/p>", html).replace('.', '')),
            int(regex_lookup("Vanaf €&nbsp;.*?tot en met €&nbsp;(.*?)<\/p>", html).replace('.', ''))
        ]
        self.vermogensbelasting_rendement = [
            round(float(regex_lookup_nogroup("<p>(\d*,\d*)%<\/p>", html)[0].replace(',', '.')) / 100, 5),
            round(float(regex_lookup_nogroup("<p>(\d*,\d*)%<\/p>", html)[1].replace(',', '.')) / 100, 5),
            round(float(regex_lookup_nogroup("<p>(\d*,\d*)%<\/p>", html)[2].replace(',', '.')) / 100, 5)
        ]

        # Heffingsvrij vermogen
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/vermogen_en_aanmerkelijk_belang/vermogen/belasting_betalen_over_uw_vermogen/heffingsvrij_vermogen/heffingsvrij_vermogen"
        html = read_html(url)

        self.heffingsvrijvermogen = regex_lookup_nogroup("<span>€ (.*?)<\/span>", html)[0]
        self.heffingsvrijvermogen = int((self.heffingsvrijvermogen * (self.fiscale_partner + 1)).replace('.', ''))

        # CO2
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/auto_en_vervoer/belastingen_op_auto_en_motor/bpm/bpm_berekenen_en_betalen/bpm_tarief/bpm-tarief-personenauto"
        html = read_html(url)

        uitstoot_min = regex_lookup_nogroup("row\">\s*<p>(\d*)\s", html)
        uitstoot_bedragen = regex_lookup_nogroup(">€&nbsp;(.*?)<\/p>", html)
        self.CO2_BPM = [
            [uitstoot_min[0], uitstoot_bedragen[0], uitstoot_bedragen[1]],
            [uitstoot_min[1], uitstoot_bedragen[2], uitstoot_bedragen[3]],
            [uitstoot_min[2], uitstoot_bedragen[4], uitstoot_bedragen[5]],
            [uitstoot_min[3], uitstoot_bedragen[6], uitstoot_bedragen[7]],
            [uitstoot_min[4], uitstoot_bedragen[8], uitstoot_bedragen[9]]
        ]
        self.CO2_BPM = [[val.replace('.', '') for val in row] for row in self.CO2_BPM]  # replace dots
        self.CO2_BPM = [[int(val) for val in row] for row in self.CO2_BPM]  # convert to int
        self.CO2_diesel_grens = int(regex_lookup("(\d*)&nbsp;gram\/km\.<\/p>", html))
        self.CO2_diesel_toeslag = float(regex_lookup("van €&nbsp;(.*?)per gram", html).replace(',', '.'))

        # Algemene heffingskorting
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/inkomstenbelasting/heffingskortingen_boxen_tarieven/heffingskortingen/algemene_heffingskorting/tabel-algemene-heffingskorting-2019"
        html = read_html(url)

        alg_korting_values = regex_lookup_nogroup("€&nbsp;(.*?)<\/p>", html)
        alg_korting_values = [alg_korting_values[val].replace('.', '') for val in range(len(alg_korting_values))]
        self.alg_korting = [
            [alg_korting_values[0], alg_korting_values[2]],
            [alg_korting_values[1], float(regex_lookup_nogroup("<\/span>-(.*?)% x", html)[0].replace(',', '.')) / 100],
            [alg_korting_values[6], alg_korting_values[7]]
        ]
        self.alg_korting_aow = [
            [alg_korting_values[8], alg_korting_values[10]],
            [alg_korting_values[1], float(regex_lookup_nogroup("<\/span>-(.*?)% x", html)[1].replace(',', '.')) / 100],
            [alg_korting_values[14], alg_korting_values[15]]
        ]
        for row in range(len(self.alg_korting)):  # Convert all compatible numbers (not floats) to int
            for val in range(len(self.alg_korting[row])):
                if type(self.alg_korting[row][val]) == str:
                    self.alg_korting[row][val] = int(self.alg_korting[row][val])
                if type(self.alg_korting_aow[row][val]) == str:
                    self.alg_korting_aow[row][val] = int(self.alg_korting_aow[row][val])

        self.alg_korting[1][1] = self.alg_korting[0][1] - self.alg_korting[1][1] * (
                self.persoon.bruto_loon_jr - self.alg_korting[1][0])
        self.alg_korting_aow[1][1] = self.alg_korting_aow[0][1] - self.alg_korting_aow[1][1] * (
                self.persoon.bruto_loon_jr - self.alg_korting_aow[1][0])

        # Arbeidskorting
        url = "https://www.belastingdienst.nl/wps/wcm/connect/bldcontentnl/belastingdienst/prive/inkomstenbelasting/heffingskortingen_boxen_tarieven/heffingskortingen/arbeidskorting/tabel-arbeidskorting-2019"
        html = read_html(url)

        arbeidskorting_schaal_values = regex_lookup_nogroup("<p>€&nbsp;(?:<span>)?(.*?)(?:<\/span>)?<\/p>", html)
        self.arbeidskorting_schaal = [
            int(arbeidskorting_schaal_values[0].replace('.', '')),
            int(arbeidskorting_schaal_values[2].replace('.', '')),
            int(arbeidskorting_schaal_values[5].replace('.', '')),
            int(arbeidskorting_schaal_values[9].replace('.', ''))
        ]
        self.arbeidskorting_korting = [
            0.01754 * self.persoon.bruto_loon_jr,
            170 + 0.28712 * (self.persoon.bruto_loon_jr - self.arbeidskorting_schaal[1]),
            3399,
            3399 - 0.06 * (self.persoon.bruto_loon_jr - self.arbeidskorting_schaal[3]),
            0
        ]
        self.arbeidskorting_korting_aow = [
            0.00898 * self.persoon.bruto_loon_jr,
            88 + 0.14689 * (self.persoon.bruto_loon_jr - self.arbeidskorting_schaal[1]),
            1740,
            1740 - 0.03069 * (self.persoon.bruto_loon_jr - self.arbeidskorting_schaal[3]),
            0
        ]


class Calculation:

    def __init__(self, persoon, auto, belasting):
        self.persoon = persoon
        self.auto = auto
        self.belasting = belasting

    def get_BTW(self):
        """
        BTW wordt als volgt berekend:
        uitgaven met lage btw (voedsel etc) * 9%   +
        uitgaven met hogte btw (incidenteel, wasmachine etc) * 21%    +
        aantal gereden kilometers per jaar * verbruik (liter per km) * btw op benzine

        Uitkomstvariabelen zijn:
        self.BTW
        """
        self.btw_laag = self.persoon.uitgaven_laag * 0.09
        self.btw_hoog = self.persoon.uitgaven_hoog * 0.21 + self.auto.km_jaar * self.auto.verbruik * self.belasting.benzine_btw
        self.btw = self.btw_hoog + self.btw_laag
        print('Calculated BTW')

    # Roken / Alcohol
    # self.rokenalcohol = 52*(self.persoon.roken*self.belasting.roken_accijns*6.50 + self.persoon.alcohol*self.belasting.alcohol_accijns*0.3)*1.21
    print('Calculated roken/alcohol')

    def get_auto(self):
        """
        Brandstof accijns is verbruik * kilometers * accijns
        BPM wordt berekend per CO2 uitstoot met verschillende schalen
        Als je een diesel rijd die meer dan 61 g/km CO2 uitstoot betaal je extra dieseltoeslag

        Uitkomstvariabelen zijn:
        self.BPM
        """
        if self.auto.brandstof.lower() == 'benzine':
            self.brandstof_accijns = self.auto.km_jaar * self.auto.verbruik * self.belasting.benzine_accijns
        elif self.auto.brandstof.lower() == 'diesel':
            self.brandstof_accijns = self.auto.km_jaar * self.auto.verbruik * self.belasting.diesel_accijns

        if self.auto.CO2_uitstoot == 0:
            self.BPM = 0
        else:
            row = find_row(self.belasting.CO2_BPM, self.auto.CO2_uitstoot)
            self.BPM = (self.auto.CO2_uitstoot - self.belasting.CO2_BPM[row][0]) * self.belasting.CO2_BPM[row][2] + \
                       self.belasting.CO2_BPM[row][1]

        if self.auto.CO2_uitstoot > self.belasting.CO2_diesel_grens and self.auto.brandstof.lower() == 'diesel':
            self.BPM = self.BPM + self.belasting.CO2_diesel_toeslag * (
                    self.auto.CO2_uitstoot - self.belasting.CO2_diesel_grens)
        print('Calculated car')

    # Loon
    def get_loon(self):
        """
        Verschillende aspecten van de loon worden hier berekend en aanvankelijk opgeslaan als losse variable, dus niet
        gekoppeld aan self (om de leesbaarheid van de berekeningen te behouden)

        Uitkomstvariabelen zijn:
        self.loontaks = loontaks
        self.loontaks_speciaal = loontaks_speciaal
        self.loontaks_aow = loontaks_aow
        self.heffingskorting = heffingskorting
        self.arbeidskorting = arbeidskorting
        self.vermogensbelasting = vermogensbelasting
        self.premievolk = premievolk
        """
        if self.persoon.leeftijd > 67:
            procent = self.belasting.loonbelasting_aow
            alg_korting = self.belasting.alg_korting_aow
            arbeid = self.belasting.arbeidskorting_korting_aow
        else:
            procent = self.belasting.loonbelasting
            alg_korting = self.belasting.alg_korting
            arbeid = self.belasting.arbeidskorting_korting

        loon_bruto = self.persoon.bruto_loon_jr
        loon_totaal = self.persoon.loon_totaal
        loon_speciaal = self.persoon.loon_speciaal
        schaal_loon = self.belasting.loonbelasting_schaal
        schaal_arbeid = self.belasting.arbeidskorting_schaal

        row = find_row(schaal_loon, loon_bruto)
        over = loon_bruto - schaal_loon[row]
        loontaks = 0
        loontaks_aow = 0
        if row == 0:
            loontaks += (schaal_loon[row + 1] - schaal_loon[row]) * procent[row]
            loontaks_aow += (schaal_loon[row + 1] - schaal_loon[row]) * self.belasting.loonbelasting_aow[row]
        else:
            for schijf in range(row):
                loontaks += ((schaal_loon[schijf + 1] - schaal_loon[schijf]) * procent[schijf])
                loontaks_aow += (
                        (schaal_loon[schijf + 1] - schaal_loon[schijf]) * self.belasting.loonbelasting_aow[schijf])
        loontaks += over * procent[row]
        loontaks_aow += over * self.belasting.loonbelasting_aow[row]

        row = find_row(schaal_loon, loon_totaal)
        loontaks_speciaal = procent[row] * loon_speciaal

        row = find_row(alg_korting, loon_bruto)
        heffingskorting = alg_korting[row][1]

        row = find_row(schaal_arbeid, loon_bruto)
        arbeidskorting = arbeid[row]

        # Vermogensbelasting
        vermogen = self.persoon.spaargeld - self.persoon.schulden - self.belasting.heffingsvrijvermogen
        vermogensbelasting = 0
        if vermogen > 0:
            row = find_row(self.belasting.vermogensbelasting_schaal, vermogen)
            over = vermogen - self.belasting.vermogensbelasting_schaal[row]
            if row == 0:
                vermogensbelasting += self.belasting.vermogensbelasting_schaal[row] * \
                                      self.belasting.vermogensbelasting_rendement[row]
            else:
                for schijf in range(row):
                    vermogensbelasting += (self.belasting.vermogensbelasting_schaal[schijf + 1] -
                                           self.belasting.vermogensbelasting_schaal[schijf]) * \
                                          self.belasting.vermogensbelasting_rendement[schijf]
                vermogensbelasting += over * self.belasting.vermogensbelasting_rendement[row]

        if self.persoon.leeftijd > 67:
            premievolk = 0
        else:
            premievolk = loontaks - (heffingskorting + arbeidskorting) - loontaks_aow

        self.loontaks = round(loontaks, 2)
        self.loontaks_speciaal = round(loontaks_speciaal, 2)
        self.loontaks_aow = round(loontaks_aow, 2)
        self.heffingskorting = round(heffingskorting, 2)
        self.arbeidskorting = round(arbeidskorting, 2)
        self.vermogensbelasting = round(vermogensbelasting, 2)
        self.premievolk = round(premievolk, 2)
        self.inkomstenbelasting = self.loontaks - (self.heffingskorting + self.arbeidskorting) - self.premievolk
        print('Calculated loon')

    # Figure
    def show(self, chart_format):
        while True:
            if chart_format.lower() != 'relative':
                if chart_format.lower() != 'absolute':
                    if chart_format.lower() != 'both':
                        chart_format = input(
                            'Please give input \'relative\' or \'absolute\' or \'both\' for chart format >>>')
                        continue
                    else:
                        break
                else:
                    break
            else:
                break
        self.taks_alles = self.inkomstenbelasting + self.premievolk + self.btw + self.belasting.rioolheffing + \
                          self.belasting.afvalheffing + self.belasting.OZB + self.brandstof_accijns + \
                          self.belasting.MRB + self.loontaks_speciaal + self.vermogensbelasting
        self.taks_pie_array = [self.inkomstenbelasting, self.premievolk, self.btw_laag, self.btw_hoog,
                               self.belasting.rioolheffing, self.belasting.afvalheffing, self.belasting.OZB,
                               self.brandstof_accijns, self.belasting.MRB, self.loontaks_speciaal,
                               self.vermogensbelasting]
        self.taks_pie_labels = ['Inkomensbelasting - kortingen', 'Premie volksverzekeringen', 'BTW laag', 'BTW hoog',
                                'Rioolheffing', 'Afvalstoffenheffing', 'OZB (Indirect)', 'Accijns op brandstof',
                                'Motorrijtuigen', 'Belasting op bonussen en vakantiegeld', 'Vermogensbelasting']
        self.taks_dict = {}
        for i in range(len(self.taks_pie_array)):
            self.taks_dict[self.taks_pie_labels[i]] = self.taks_pie_array[i]
        self.taks_dict_sort = sorted(self.taks_dict.items(), key=lambda x: x[1])  # sort descending

        # Chance list back to dict
        self.taks_dict = {}
        for i in range(len(self.taks_dict_sort)):
            self.taks_dict[self.taks_dict_sort[i][0]] = self.taks_dict_sort[i][1]

        self.taks_ratio = self.taks_alles / self.persoon.loon_totaal

        print('Made data arrays.\nGenerating pie plot...')

        if chart_format.lower() == 'relative':
            chart_format = '%1.1f%%'
        elif chart_format.lower() == 'absolute':
            chart_format = lambda p: '{:.0f}'.format((p * self.taks_alles) / 100)
        else:  # show both
            chart_format = lambda p: '{:.0f} ({:.2f}%)'.format((p * self.taks_alles) / 100, p)

        fig1, ax1 = plt.subplots()
        ax1.pie(list(self.taks_dict.values()), labels=list(self.taks_dict.keys()), autopct=chart_format)
        ax1.axis('equal')
        title_string = ('Totale jaarlijkse belasting: {} euro = {}%'.format(round(self.taks_alles, 2),
                                                                            round(self.taks_ratio * 100, 2)))
        plt.title(title_string)
        plt.show()

        print('Finished')


def regex_lookup(regex_string, data_to_search):
    """
    Functie waarmee je een string uit tekst kan halen waarmee maar 1 match is

    :param regex_string:    De regex code waarmee je wil zoeken
    :param data_to_search:  De data waarin je wil zoeken
    :return:                De tekst waar je naar zocht (als alles goed is gegaan)
    """
    reg = re.compile(r"" + regex_string + "", re.DOTALL)
    data = reg.search(data_to_search)
    return data.group(1) if data else '0'


def regex_lookup_nogroup(regex_string, data_to_search):
    """
    Functie waarmee je een string uit tekst kan halen die meerdere matches heeft
    Selecteer de match die je wil mbv regex_lookup_nogroup(*args)[nummer match]

    :param regex_string:    De regex code waarmee je wil zoeken
    :param data_to_search:  De data waarin je wil zoeken
    :return:                De tekst waar je naar zocht (als alles goed is gegaan)
    """
    reg = re.compile(r"" + regex_string + "", re.DOTALL)
    data = reg.findall(data_to_search)
    return data


def read_html(url):
    """
    Convert een url naar een html tekstbestand van die pagina

    :param url:     Een url
    :return:        De html code in plaintext van die url
    """
    response = urllib.request.urlopen(url)
    html = response.read().decode(response.headers.get_content_charset())
    return html


def find_row(table, var):
    """
    Vind de juiste rij van een schaaltabel waarin een persoon valt
    Voorbeeld:  je verdient 20000 euro, schaal 1 van de tabel gaat tot 15000, schaal 2 tot 30000
                je valt dan in schaal 1 (= rij 0)

    :param table:   Tabel met de cutoffs van de verschillende schijven in kolom 0
    :param var:     Hoogte van de waarde die je zoekt (bv loon)
    :return:        Rijnummer waarin je valt (beginnend vanaf 0)
    """
    for row in range(len(table) - 1, -1, -1):  # count down from rows in table to 0
        try:
            if var > table[row][0]:
                return row
        except TypeError:
            if var > table[row]:
                return row


if __name__ == '__main__':
    boris = Persoon('5531vg', 25, 2500, 10000, 300, 600, 2000, 35000, 1500, 3000, 93)
    auto = Voertuig(boris, '85tdpv', 2500, 20000)
    tax = Belasting(boris, auto, 2, 0)
    resultaat = Calculation(boris, auto, tax)
    resultaat.get_auto()
    resultaat.get_loon()
    resultaat.get_BTW()
    resultaat.show('both')